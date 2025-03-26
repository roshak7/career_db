from datetime import datetime

import openpyxl

from career.models import Persons, Career_structure
from career.upload.generator_id import id_generate
from career.upload.matrix import matrix


def parsing_xlsx_file_multi(file, request):
    try:
        cols = 0
        rows = 7

        new_org_structure = []
        new_org_successor = []
        in_file_xlsx = file  # 'E:\\dev\\career-backend\\111.xlsx'
        wb1 = openpyxl.load_workbook(in_file_xlsx)
        sh = wb1.active
        id_ruk_main = ''
        s_row = 0
        pp = 0


        for x in range(rows, 205):
            # if sh.cell(row=x, column=2).value == 'Защищенность':
            if str(sh.cell(row=x, column=3).value).lower() == 'защищенность' or str(
                    sh.cell(row=x, column=3).value).lower() == 'защищённость':
                s_row = x + 1
                break
        for x in range(rows, 205):
            if sh.cell(row=x, column=1).value == None and sh.cell(row=x, column=2).value == None and sh.cell(row=x,
                                                                                                             column=3).value == None:
                pass
            else:
                dd = [str(sh.cell(row=x, column=1).value), str(sh.cell(row=x, column=2).value),
                      str(sh.cell(row=x, column=3).value)]
                pp += 1
                if pp == 1:
                    name_card = max(dd)
                if pp == 2:
                    name_company = max(dd)
                    break
        ruk_dic = {}
        for x in range(s_row, 205):
            try:
                if int(sh.cell(row=x, column=1).value)>0:

                    position_full_name = str(sh.cell(row=x, column=11 + cols).value).strip()

                    person_fio = str(sh.cell(row=x, column=12 + cols).value).strip()
                    person_fio = " ".join(person_fio.split())
                    id_ruk = id_generate(person_fio, position_full_name)
                    ruk_dic.update({str(sh.cell(row=x, column=1).value):id_ruk})
            except:
                pass

        for x in range(s_row, 205):

            if sh.cell(row=x, column=3 + cols).value == None or str(
                    sh.cell(row=x, column=3 + cols).value).strip() == '' or len(
                str(sh.cell(row=x, column=3 + cols).value)) <= 2 or len(
                str(sh.cell(row=x, column=4 + cols).value)) <= 2 or len(
                str(sh.cell(row=x, column=3 + cols).value).strip()) <= 2:
                continue

            # id_ruk.replace("00:00:00", "")
            # department_full_name = str(sh.cell(row=x, column=10 + cols).value)
            protection = str(sh.cell(row=x, column=3 + cols).value).strip()
            organization_name = str(sh.cell(row=x, column=4 + cols).value).strip()
            parent = str(sh.cell(row=x, column=2).value)
            age = str(sh.cell(row=x, column=5 + cols).value)
            birthday = sh.cell(row=x, column=6 + cols).value
            if birthday != None:
                if type(birthday) != datetime.date:
                    try:
                        birthday = datetime.strptime(birthday, '%d.%m.%Y')
                    except:
                        pass
            else:
                birthday = None
            if len(age) > 3:
                age = '0'
            scientific_degree = str(sh.cell(row=x, column=7 + cols).value).replace("None", "").replace("-", "")
            personnel_reserve = str(sh.cell(row=x, column=8 + cols).value).replace("None", "").replace("Нет",
                                                                                                       "").replace(
                "-", "")
            personnel_reserve_dop = str(sh.cell(row=x, column=9 + cols).value).replace("None", "").replace("Нет", "")

            position_critical = str(sh.cell(row=x, column=10 + cols).value)

            position_full_name = str(sh.cell(row=x, column=11 + cols).value).strip()

            if sh.cell(row=x, column=12 + cols).value==None or str(sh.cell(row=x, column=12 + cols).value).strip()=='':
                person_fio = 'Вакансия'
            else:
                person_fio = str(sh.cell(row=x, column=12 + cols).value).strip()
                person_fio = " ".join(person_fio.split())

            id_ruk = id_generate(person_fio, position_full_name)
            person_id_photo = id_ruk

            contract_end_date = sh.cell(row=x, column=13 + cols).value

            if type(contract_end_date) != datetime:
                try:
                    contract_end_date = datetime.strptime(contract_end_date, '%d.%m.%Y')
                except:
                    contract_end_date = None

            obj, created = Persons.objects.update_or_create(
                person_id=id_ruk,
                defaults={'age': age,
                          'protection': protection,
                          'person_id_photo': person_id_photo,
                          'position_critical': position_critical,
                          # 'department_full_name': department_full_name,
                          'person_name': person_fio,
                          'person_fio': person_fio,
                          'organization_name': organization_name,
                          'birthday': birthday,
                          'position_short_name': position_full_name,
                          'position_full_name': position_full_name,
                          'scientific_degree': scientific_degree,
                          'personnel_reserve': personnel_reserve,
                          'personnel_reserve_dop': personnel_reserve_dop,
                          'contract_end_date': contract_end_date

                          },
            )
            if parent=='0':
                id_ruk_main=''
            else:
                try:
                    id_ruk_main=ruk_dic[parent]
                except:
                    id_ruk_main=''
                    pass
            new_org_structure.append({'id': id_ruk, 'img': '/api/v1/img/%s/' % person_id_photo, 'parentID': id_ruk_main})
            # if id_ruk_main == '':
            #     id_ruk_main = id_ruk
            if sh.cell(row=x, column=15 + cols).value == None:
                continue
            for xx in range(0, 8):

                # id_ruk.replace("00:00:00", "")
                # c 14 по 27

                # department_full_name = str(sh.cell(row=x + xx, column=14 + cols).value)
                position_full_name = str(sh.cell(row=x + xx, column=14 + cols).value)
                person_fio = str(sh.cell(row=x + xx, column=15 + cols).value).strip()
                person_fio = " ".join(person_fio.split())
                assessment_potential = str(sh.cell(row=x + xx, column=16 + cols).value)
                assessment_potential_year = str(sh.cell(row=x + xx, column=17 + cols).value)
                if len(assessment_potential_year)>4:
                    assessment_potential_year=assessment_potential_year[-3:]
                # organization_name берем из переменной руководителя
                age = str(sh.cell(row=x, column=18 + cols).value)
                if len(age) > 3:
                    age = '0'
                birthday = sh.cell(row=x + xx, column=19 + cols).value
                if type(birthday) != datetime.date:
                    try:
                        birthday = datetime.strptime(birthday, '%d.%m.%Y')
                    except:
                        pass
                readiness = str(sh.cell(row=x + xx, column=20 + cols).value)
                mobility = str(sh.cell(row=x + xx, column=21 + cols).value)
                scientific_degree = str(sh.cell(row=x + xx, column=22 + cols).value).replace("None", "").replace("-",
                                                                                                                 "")
                personnel_reserve = str(sh.cell(row=x + xx, column=23 + cols).value).replace("None", "").replace("-",
                                                                                                                 "")
                personnel_reserve_dop = str(sh.cell(row=x + xx, column=24 + cols).value).replace("None", "").replace(
                    "-",
                    "")

                id_suc = id_generate(person_fio, position_full_name)
                person_id_photo = id_suc
                id_suc = id_suc + "_" + id_ruk

                kpi_3 = sh.cell(row=x + xx, column=25 + cols).value
                kpi_2 = sh.cell(row=x + xx, column=26 + cols).value
                kpi_1 = sh.cell(row=x + xx, column=27 + cols).value
                try:
                    if len(str(kpi_1)) < 2:
                        kpi_1 = None
                    if len(str(kpi_2)) < 2:
                        kpi_2 = None
                    if len(str(kpi_3)) < 2:
                        kpi_3 = None
                except:
                    pass
                # profile_str = str(sh.cell(row=x + xx, column=26 + cols).value)
                # profile = True if profile_str == 'Да' else False
                # participation_assessment_str = str(sh.cell(row=x + xx, column=27 + cols).value)
                # participation_assessment = True if participation_assessment_str == 'Да' else False

                obj, created = Persons.objects.update_or_create(
                    person_id=id_suc,
                    defaults={'readiness': readiness,
                              'age': age,
                              'mobility': mobility,
                              'personnel_reserve_dop': personnel_reserve_dop,
                              'person_id_photo': person_id_photo,
                              # 'department_full_name': department_full_name,
                              'person_name': person_fio,
                              'person_fio': person_fio,
                              'assessment_potential': assessment_potential,
                              'assessment_potential_year': assessment_potential_year,
                              'organization_name': organization_name,
                              'birthday': birthday,
                              'position_short_name': position_full_name,
                              'position_full_name': position_full_name,
                              'scientific_degree': scientific_degree,
                              'personnel_reserve': personnel_reserve,
                              'kpi_3': kpi_3,
                              'kpi_2': kpi_2,
                              'kpi_1': kpi_1
                              # 'profile': profile,
                              # 'participation_assessment': participation_assessment
                              },
                )
                new_org_successor.append(
                    {'id': id_suc, 'img': '/api/v1/img/%s/' % person_id_photo, 'parentID': id_ruk, 'protect': 3})

                if sh.cell(row=x + xx + 1, column=12 + cols).value != None:
                    break
                else:
                    if sh.cell(row=x + xx + 1, column=19 + cols).value == None:
                        break
        if request.user.username:
            user_l = request.user
        else:
            user_l = None
        org_card = Career_structure.objects.create(jdata_org_structure=new_org_structure,
                                                   jdata_org_successor=new_org_successor, name=name_card,
                                                   company=name_company, id_user=user_l, type_map='multilevel')
        org_card.save()
        # try:
        #     Persons.objects.get(person_id=id_ruk)
        # except:
        #     person = Persons(person_id=id_ruk)
        # for y in range(2, 26):
        #     print(sh.cell(row=x, column=y).value)
        # print("----")
        # matrix(id=org_card.pk)
    except Exception as e:
        print(e.messages[0])
        return [0, {'errMsg': e.messages[0]}]
    return [org_card.pk, {'report': 'Карта загружена'}]
def parsing_xlsx_file_2(file, request):
    try:
        cols = -1
        rows = 7

        new_org_structure = []
        new_org_successor = []
        # name_struct = "Структура от %s" % str(datetime.datetime.now())
        in_file_xlsx = file  # 'E:\\dev\\career-backend\\111.xlsx'
        wb1 = openpyxl.load_workbook(in_file_xlsx)
        sh = wb1.active
        # name_card = str(sh.cell(row=6, column=3 + cols).value).strip()
        # name_company = str(sh.cell(row=8, column=3 + cols).value).strip()
        # name_struct = name_card  # + ' ' + str(datetime.datetime.now())

        id_ruk_main = ''
        s_row = 0
        pp = 0
        for x in range(rows, 205):
            if str(sh.cell(row=x, column=2).value).lower() == '0':
                return parsing_xlsx_file_multi(file, request)

            # if sh.cell(row=x, column=2).value == 'Защищенность':
            if str(sh.cell(row=x, column=2).value).lower() == 'защищенность' or str(
                    sh.cell(row=x, column=2).value).lower() == 'защищённость':
                s_row = x + 1
                break
        for x in range(rows, 205):
            if sh.cell(row=x, column=1).value == None and sh.cell(row=x, column=2).value == None and sh.cell(row=x,
                                                                                                             column=3).value == None:
                pass
            else:
                # dd = [str(sh.cell(row=x, column=1).value), str(sh.cell(row=x, column=2).value),
                #       str(sh.cell(row=x, column=3).value)]
                dds = [str(sh.cell(row=x, column=1).value).replace('None',''), str(sh.cell(row=x, column=2).value).replace('None',''),
                      str(sh.cell(row=x, column=3).value).replace('None','')]
                pp += 1
                if pp == 1:
                    name_card = max(dds, key=len)
                if pp == 2:

                    name_company = max(dds, key=len)
                    break

        for x in range(s_row, 205):

            if sh.cell(row=x, column=3 + cols).value == None or str(
                    sh.cell(row=x, column=3 + cols).value).strip() == '' or len(
                str(sh.cell(row=x, column=3 + cols).value)) <= 2 or len(
                str(sh.cell(row=x, column=4 + cols).value)) <= 2 or len(
                str(sh.cell(row=x, column=3 + cols).value).strip()) <= 2:
                continue

            # id_ruk.replace("00:00:00", "")
            # department_full_name = str(sh.cell(row=x, column=10 + cols).value)
            protection = str(sh.cell(row=x, column=3 + cols).value).strip()
            organization_name = str(sh.cell(row=x, column=4 + cols).value).strip()

            age = str(sh.cell(row=x, column=5 + cols).value)
            birthday = sh.cell(row=x, column=6 + cols).value
            if birthday != None:
                if type(birthday) != datetime.date:
                    try:
                        birthday = datetime.strptime(birthday, '%d.%m.%Y')
                    except:
                        pass
            else:
                birthday = None
            if len(age) > 3:
                age = '0'
            scientific_degree = str(sh.cell(row=x, column=7 + cols).value).replace("None", "").replace("-", "")
            personnel_reserve = str(sh.cell(row=x, column=8 + cols).value).replace("None", "").replace("Нет",
                                                                                                       "").replace(
                "-", "")
            personnel_reserve_dop = str(sh.cell(row=x, column=9 + cols).value).replace("None", "").replace("Нет", "")

            position_critical = str(sh.cell(row=x, column=10 + cols).value)
            # position_full_name = str(sh.cell(row=x, column=11 + cols).value).strip()
            position_full_name = str(sh.cell(row=x, column=11 + cols).value).strip()

            if sh.cell(row=x, column=12 + cols).value==None or str(sh.cell(row=x, column=12 + cols).value).strip()=='':
                person_fio = 'Вакансия'
            else:
                person_fio = str(sh.cell(row=x, column=12 + cols).value).strip()
                person_fio = " ".join(person_fio.split())


            id_ruk = id_generate(person_fio, position_full_name)
            person_id_photo = id_ruk

            contract_end_date = sh.cell(row=x, column=13 + cols).value

            if type(contract_end_date) != datetime:
                try:
                    contract_end_date = datetime.strptime(contract_end_date, '%d.%m.%Y')
                except:
                    contract_end_date = None

            obj, created = Persons.objects.update_or_create(
                person_id=id_ruk,
                defaults={'age': age,
                          'protection': protection,
                          'person_id_photo': person_id_photo,
                          'position_critical': position_critical,
                          # 'department_full_name': department_full_name,
                          'person_name': person_fio,
                          'person_fio': person_fio,
                          'organization_name': organization_name,
                          'birthday': birthday,
                          'position_short_name': position_full_name,
                          'position_full_name': position_full_name,
                          'scientific_degree': scientific_degree,
                          'personnel_reserve': personnel_reserve,
                          'personnel_reserve_dop': personnel_reserve_dop,
                          'contract_end_date': contract_end_date

                          },
            )

            new_org_structure.append({'id': id_ruk, 'img': '/api/v1/img/%s/' % person_id_photo, 'parentID': id_ruk_main})
            if id_ruk_main == '':
                id_ruk_main = id_ruk
            if sh.cell(row=x, column=15 + cols).value == None:
                continue
            for xx in range(0, 8):

                # id_ruk.replace("00:00:00", "")
                # c 14 по 27

                # department_full_name = str(sh.cell(row=x + xx, column=14 + cols).value)
                position_full_name = str(sh.cell(row=x + xx, column=14 + cols).value)
                person_fio = str(sh.cell(row=x + xx, column=15 + cols).value).strip()
                person_fio = " ".join(person_fio.split())
                assessment_potential = str(sh.cell(row=x + xx, column=16 + cols).value)
                assessment_potential_year = str(sh.cell(row=x + xx, column=17 + cols).value)
                if len(assessment_potential_year)>4:
                    assessment_potential_year=assessment_potential_year[-3:]
                # organization_name берем из переменной руководителя
                age = str(sh.cell(row=x, column=18 + cols).value)
                if len(age) > 3:
                    age = '0'
                birthday = sh.cell(row=x + xx, column=19 + cols).value
                if type(birthday) != datetime.date:
                    try:
                        birthday = datetime.strptime(birthday, '%d.%m.%Y')
                    except:
                        pass
                readiness = str(sh.cell(row=x + xx, column=20 + cols).value)
                mobility = str(sh.cell(row=x + xx, column=21 + cols).value)
                scientific_degree = str(sh.cell(row=x + xx, column=22 + cols).value).replace("None", "").replace("-",
                                                                                                                 "")
                personnel_reserve = str(sh.cell(row=x + xx, column=23 + cols).value).replace("None", "").replace("-",
                                                                                                                 "")
                personnel_reserve_dop = str(sh.cell(row=x + xx, column=24 + cols).value).replace("None", "").replace(
                    "-",
                    "")

                id_suc = id_generate(person_fio, position_full_name)
                person_id_photo = id_suc
                id_suc = id_suc + "_" + id_ruk

                kpi_3 = sh.cell(row=x + xx, column=25 + cols).value
                kpi_2 = sh.cell(row=x + xx, column=26 + cols).value
                kpi_1 = sh.cell(row=x + xx, column=27 + cols).value
                try:
                    if len(str(kpi_1)) < 2:
                        kpi_1 = None
                    if len(str(kpi_2)) < 2:
                        kpi_2 = None
                    if len(str(kpi_3)) < 2:
                        kpi_3 = None
                except:
                    pass
                # profile_str = str(sh.cell(row=x + xx, column=26 + cols).value)
                # profile = True if profile_str == 'Да' else False
                # participation_assessment_str = str(sh.cell(row=x + xx, column=27 + cols).value)
                # participation_assessment = True if participation_assessment_str == 'Да' else False

                obj, created = Persons.objects.update_or_create(
                    person_id=id_suc,
                    defaults={'readiness': readiness,
                              'age': age,
                              'mobility': mobility,
                              'personnel_reserve_dop': personnel_reserve_dop,
                              'person_id_photo': person_id_photo,
                              # 'department_full_name': department_full_name,
                              'person_name': person_fio,
                              'person_fio': person_fio,
                              'assessment_potential': assessment_potential,
                              'assessment_potential_year': assessment_potential_year,
                              'organization_name': organization_name,
                              'birthday': birthday,
                              'position_short_name': position_full_name,
                              'position_full_name': position_full_name,
                              'scientific_degree': scientific_degree,
                              'personnel_reserve': personnel_reserve,
                              'kpi_3': kpi_3,
                              'kpi_2': kpi_2,
                              'kpi_1': kpi_1
                              # 'profile': profile,
                              # 'participation_assessment': participation_assessment
                              },
                )
                new_org_successor.append(
                    {'id': id_suc, 'img': '/api/v1/img/%s/' % person_id_photo, 'parentID': id_ruk, 'protect': 3})

                if sh.cell(row=x + xx + 1, column=12 + cols).value != None:
                    break
                else:
                    if sh.cell(row=x + xx + 1, column=19 + cols).value == None:
                        break
        if request.user.username:
            user_l = request.user
        else:
            user_l = None
        org_card = Career_structure.objects.create(jdata_org_structure=new_org_structure,
                                                   jdata_org_successor=new_org_successor, name=name_card,
                                                   company=name_company, id_user=user_l)
        org_card.save()
        # try:
        #     Persons.objects.get(person_id=id_ruk)
        # except:
        #     person = Persons(person_id=id_ruk)
        # for y in range(2, 26):
        #     print(sh.cell(row=x, column=y).value)
        # print("----")
        # matrix(id=org_card.pk)
    except Exception as e:
        print(e.messages[0])
        return [0, {'errMsg': e.messages[0]}]
    return [org_card.pk, {'report': 'Карта загружена'}]
